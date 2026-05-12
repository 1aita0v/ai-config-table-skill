#!/usr/bin/env python3
"""Generate a sample workbook for the walkthrough.

Creates examples/sample.xlsx with 3 sheets:
- Item: 4-row header (CN / EN / type / comment) + 5 data rows
- LocText: 2-row header + 5 data rows
- Reward: 1-row header + 3 data rows

Run from the skill root:
    python3 examples/build_sample.py
"""
from __future__ import annotations

from pathlib import Path

try:
    from openpyxl import Workbook
except ModuleNotFoundError:
    raise SystemExit("openpyxl is required. Install with: pip install openpyxl")


HERE = Path(__file__).resolve().parent
OUT = HERE / "sample.xlsx"


def build() -> None:
    wb = Workbook()

    # --- Item sheet: 4-row header ---
    item = wb.active
    item.title = "Item"
    item.append(["道具ID", "名称", "描述", "图标", "品质", "本地化键"])
    item.append(["ItemID", "Name", "Desc", "Icon", "Quality", "LocKey"])
    item.append(["int", "string", "string", "string", "int", "string"])
    item.append(["主键", "运行时文本", "运行时文本", "资源引用", "1-5", "Loc 表外键"])
    item.append([10001, "Sword", "A basic sword.", "icon_sword", 1, "ITEM_10001_NAME"])
    item.append([10002, "Shield", "A wooden shield.", "icon_shield", 1, "ITEM_10002_NAME"])
    item.append([10003, "Potion", "Restores HP.", "icon_potion", 2, "ITEM_10003_NAME"])
    item.append([10004, "Bow", "Ranged weapon.", "icon_bow", 2, "ITEM_10004_NAME"])
    item.append([10005, "Staff", "Magic weapon.", "icon_staff", 3, "ITEM_10005_NAME"])

    # --- LocText sheet: 2-row header (CN / EN field) ---
    loc = wb.create_sheet("LocText")
    loc.append(["本地化键", "中文文本", "英文文本"])
    loc.append(["LocKey", "TextCN", "TextEN"])
    loc.append(["ITEM_10001_NAME", "长剑", "Sword"])
    loc.append(["ITEM_10002_NAME", "木盾", "Shield"])
    loc.append(["ITEM_10003_NAME", "药水", "Potion"])
    loc.append(["ITEM_10004_NAME", "弓", "Bow"])
    loc.append(["ITEM_10005_NAME", "法杖", "Staff"])

    # --- Reward sheet: 1-row header ---
    reward = wb.create_sheet("Reward")
    reward.append(["RewardID", "ItemID", "Count"])
    reward.append([20001, 10001, 1])
    reward.append([20002, 10003, 5])
    reward.append([20003, 10005, 1])

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
