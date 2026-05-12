#!/usr/bin/env python3
"""Cross-table reference validation.

Read a workbook (typically a candidate produced by patch_xlsx.py) and verify
that foreign-key-like columns actually resolve to a row in their target sheet.
Reports orphan references — values that don't match any primary key in the
target.

Two modes:

1. Auto-detect (default). The script:
   - finds each sheet's "primary key" column by header heuristics
     (columns named like `Id`, `Key`, `ItemID`, `id`, `编号`, `主键`, `*_id`)
   - for each non-key column whose header ends in `Key`/`ID`/`Id`, looks for
     another sheet whose primary key header *exactly matches* the column name
   - reports any such reference whose values can't be found

   Example: a column `Item.LocKey` will be checked against `LocText.LocKey`
   if `LocKey` is `LocText`'s primary key.

2. Explicit schema. Pass `--schema schema.json` with the form:

       {
         "field_rows":  {"Item": 2, "LocText": 2},
         "meta_rows":   {"Item": [1,3,4], "LocText": [1]},
         "refs": [
           {"from_sheet": "Item", "from_field": "LocKey",
            "to_sheet": "LocText", "to_field": "LocKey"}
         ]
       }

   The schema overrides auto-detection for the listed refs but does not
   suppress the auto-detected ones unless you pass `--no-auto`.

Output: Markdown by default, JSON with `--format json`.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any


SUPPORTED = {".xlsx", ".xlsm"}

KEY_EXACT = {"id", "key", "编号", "主键", "uid"}
KEY_SUFFIX_RE = re.compile(r"_id$", re.IGNORECASE)
KEY_CAMEL_RE = re.compile(r"^[A-Za-z][A-Za-z0-9]*(ID|Id|Key)$")
FOREIGN_SUFFIX_RE = re.compile(r"(ID|Id|Key)$")
ASCII_FIELD_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]*$")
CHINESE_RE = re.compile(r"[一-鿿]")
TYPE_ANNOTATION_VALUES = {
    "int", "integer", "str", "string", "float", "double", "bool", "boolean",
    "list", "array", "dict", "object", "json", "date", "datetime", "long",
    "short", "char", "varchar", "text", "number",
}
CN_ANNOTATION_KEYWORDS = (
    "主键", "外键", "唯一", "必填", "可选", "注释", "说明", "类型", "枚举",
    "引用", "默认", "废弃", "弃用", "运行时", "客户端", "服务端",
)


def load_openpyxl():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required for reference validation.\n"
            "Install with:  pip install openpyxl"
        ) from exc
    return load_workbook


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_key_candidate(name: str) -> bool:
    if not name:
        return False
    low = name.strip().lower()
    if low in KEY_EXACT:
        return True
    if KEY_SUFFIX_RE.search(name):
        return True
    if KEY_CAMEL_RE.match(name):
        return True
    return False


def looks_like_foreign(name: str) -> bool:
    if not name:
        return False
    return bool(FOREIGN_SUFFIX_RE.search(name))


def _looks_ascii_field_row(values: list[str]) -> bool:
    nonempty = [v for v in values if v]
    if not nonempty:
        return False
    ascii_hits = sum(1 for v in nonempty if ASCII_FIELD_RE.match(v))
    return ascii_hits >= max(1, len(nonempty) * 2 // 3)


def _looks_annotation_row(values: list[str]) -> bool:
    nonempty = [v for v in values if v]
    if not nonempty:
        return False
    type_hits = sum(1 for v in nonempty if v.lower() in TYPE_ANNOTATION_VALUES)
    cn_hits = sum(1 for v in nonempty if any(kw in v for kw in CN_ANNOTATION_KEYWORDS))
    short_cn = sum(1 for v in nonempty if len(v) <= 12 and not any(c.isdigit() for c in v) and CHINESE_RE.search(v))
    return (type_hits + cn_hits) >= max(1, len(nonempty) // 2) or short_cn == len(nonempty)


def auto_detect_layout(ws) -> tuple[int, list[int]]:
    """Find the most likely field_row + meta_rows for one sheet.

    Strategy: scan the first ~8 rows; pick the row that looks most like ASCII
    field names. Rows above it become meta_rows. Rows below it that look like
    annotations / type comments are added to meta_rows too.
    """
    scan_limit = min(ws.max_row, 8)
    sample_rows: list[list[str]] = []
    for row_idx in range(1, scan_limit + 1):
        sample_rows.append([norm(ws.cell(row=row_idx, column=c).value) for c in range(1, ws.max_column + 1)])

    best_field_row = 1
    best_score = -1
    for idx, row in enumerate(sample_rows, start=1):
        if not _looks_ascii_field_row(row):
            continue
        score = sum(1 for v in row if v)
        if score > best_score:
            best_score = score
            best_field_row = idx
    meta_rows = list(range(1, best_field_row))
    # Extend meta_rows downward through annotation rows.
    next_row = best_field_row + 1
    while next_row <= len(sample_rows) and _looks_annotation_row(sample_rows[next_row - 1]):
        meta_rows.append(next_row)
        next_row += 1
    return best_field_row, sorted(meta_rows)


def read_sheet(ws, field_row: int, meta_rows: list[int]) -> dict[str, Any]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        name = norm(ws.cell(row=field_row, column=col).value)
        if name and name not in headers:
            headers[name] = col
    last_header = max([field_row] + list(meta_rows))
    data_start = last_header + 1

    columns: dict[str, list[tuple[int, str]]] = {name: [] for name in headers}
    for row_idx in range(data_start, ws.max_row + 1):
        row_has_data = False
        for name, col in headers.items():
            value = norm(ws.cell(row=row_idx, column=col).value)
            if value:
                row_has_data = True
            columns[name].append((row_idx, value))
        # keep all rows (including blanks) so row indices line up
        if not row_has_data:
            # but trim trailing empty rows at the end of the loop
            pass

    # Trim trailing empty rows across all columns.
    if columns:
        any_col = next(iter(columns.values()))
        while any_col and all(not col[-1][1] for col in columns.values()):
            for col in columns.values():
                col.pop()
            any_col = columns[next(iter(columns))]

    # Primary key: first header that looks like a key.
    primary_key = None
    for name in headers:
        if is_key_candidate(name):
            primary_key = name
            break

    return {
        "name": ws.title,
        "field_row": field_row,
        "meta_rows": list(meta_rows),
        "data_start_row": data_start,
        "headers": headers,
        "primary_key": primary_key,
        "columns": columns,
    }


def auto_detect_refs(sheets: dict[str, dict[str, Any]]) -> list[dict[str, str]]:
    refs: list[dict[str, str]] = []
    # Build target index: header-name -> [(sheet, primary_key)]
    pk_by_sheet: dict[str, str] = {
        name: info["primary_key"] for name, info in sheets.items() if info["primary_key"]
    }
    for from_sheet, info in sheets.items():
        pk = info["primary_key"]
        for field in info["headers"]:
            if field == pk:
                continue
            if not looks_like_foreign(field):
                continue
            # Try exact match against another sheet's primary key.
            for to_sheet, to_pk in pk_by_sheet.items():
                if to_sheet == from_sheet:
                    continue
                if to_pk == field:
                    refs.append(
                        {
                            "from_sheet": from_sheet,
                            "from_field": field,
                            "to_sheet": to_sheet,
                            "to_field": to_pk,
                            "source": "auto",
                        }
                    )
                    break
    return refs


def check_ref(ref: dict[str, str], sheets: dict[str, dict[str, Any]]) -> dict[str, Any]:
    from_info = sheets.get(ref["from_sheet"])
    to_info = sheets.get(ref["to_sheet"])
    if not from_info:
        return {**ref, "error": f"from_sheet '{ref['from_sheet']}' not found"}
    if not to_info:
        return {**ref, "error": f"to_sheet '{ref['to_sheet']}' not found"}
    from_col = from_info["columns"].get(ref["from_field"])
    to_col = to_info["columns"].get(ref["to_field"])
    if from_col is None:
        return {**ref, "error": f"from_field '{ref['from_field']}' not found in {ref['from_sheet']}"}
    if to_col is None:
        return {**ref, "error": f"to_field '{ref['to_field']}' not found in {ref['to_sheet']}"}
    target_keys = {value for _, value in to_col if value}
    orphans: list[dict[str, Any]] = []
    checked = 0
    for row_idx, value in from_col:
        if not value:
            continue
        checked += 1
        if value not in target_keys:
            orphans.append({"row": row_idx, "value": value})
    return {
        **ref,
        "checked": checked,
        "orphan_count": len(orphans),
        "orphans": orphans[:50],
        "truncated": len(orphans) > 50,
    }


def load_schema(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {
        "field_rows": payload.get("field_rows", {}),
        "meta_rows": {k: list(v) for k, v in payload.get("meta_rows", {}).items()},
        "refs": payload.get("refs", []),
    }


def render_md(report: dict[str, Any]) -> str:
    lines = ["# Cross-Table Reference Validation", "", f"- Workbook: `{report['workbook']}`", ""]
    sheets = report["sheets"]
    lines.append(f"## Sheets scanned ({len(sheets)})")
    lines.append("")
    lines.append("| Sheet | Primary Key | Field Row | Data Start | Rows |")
    lines.append("|---|---|---:|---:|---:|")
    for info in sheets:
        rows_observed = len(next(iter(info.get("columns", {}).values()), []))
        lines.append(
            f"| {info['name']} | {info.get('primary_key') or '—'} | "
            f"{info['field_row']} | {info['data_start_row']} | {rows_observed} |"
        )
    lines.append("")

    refs = report["refs"]
    if not refs:
        lines.append("## References")
        lines.append("")
        lines.append("No cross-table references detected (and none provided via --schema).")
        return "\n".join(lines)

    total_orphans = sum(r.get("orphan_count", 0) for r in refs)
    if total_orphans == 0:
        lines.append(f"## ✅ All {len(refs)} reference(s) resolve cleanly")
    else:
        lines.append(f"## ❌ {total_orphans} orphan value(s) across {len(refs)} reference(s)")
    lines.append("")

    for ref in refs:
        title = f"### {ref['from_sheet']}.{ref['from_field']} → {ref['to_sheet']}.{ref['to_field']}"
        if ref.get("source") == "auto":
            title += "  *(auto-detected)*"
        lines.append(title)
        lines.append("")
        if ref.get("error"):
            lines.append(f"- **ERROR**: {ref['error']}")
            lines.append("")
            continue
        lines.append(f"- Checked: {ref.get('checked', 0)} non-empty value(s)")
        if ref.get("orphan_count", 0) == 0:
            lines.append("- Orphans: none ✅")
            lines.append("")
            continue
        lines.append(f"- Orphans: **{ref['orphan_count']}**")
        if ref.get("truncated"):
            lines.append(f"- (showing first 50; total = {ref['orphan_count']})")
        lines.append("")
        lines.append("| Row | Value | Note |")
        lines.append("|---:|---|---|")
        for orphan in ref["orphans"]:
            value = orphan["value"].replace("|", "\\|")
            lines.append(f"| {orphan['row']} | {value} | not found in {ref['to_sheet']}.{ref['to_field']} |")
        lines.append("")
    return "\n".join(lines)


def parse_meta_rows(value: str) -> list[int]:
    if not value:
        return []
    parts = [p.strip() for p in value.split(",") if p.strip()]
    try:
        return [int(p) for p in parts]
    except ValueError as exc:
        raise SystemExit(f"--meta-rows expects comma-separated integers, got: {value}") from exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate cross-table references in a workbook.")
    parser.add_argument("--workbook", type=Path, required=True, help="Path to .xlsx / .xlsm to validate.")
    parser.add_argument(
        "--field-row", type=int, default=None,
        help="Default 1-based field row applied to all sheets. Omit to auto-detect per sheet "
             "(ASCII-field-name row + annotation rows above/below).",
    )
    parser.add_argument(
        "--meta-rows", type=str, default="",
        help="Default comma-separated meta rows applied to all sheets.",
    )
    parser.add_argument(
        "--schema", type=Path, default=None,
        help="Optional JSON schema with per-sheet field_rows/meta_rows and explicit refs.",
    )
    parser.add_argument(
        "--no-auto", action="store_true",
        help="Skip auto-detection of refs; only use refs declared in --schema.",
    )
    parser.add_argument("--output", type=Path, help="Write the report to a file.")
    parser.add_argument("--format", choices=("md", "json"), default="md")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")
    if args.workbook.suffix.lower() not in SUPPORTED:
        raise SystemExit(f"Unsupported workbook type: {args.workbook.suffix}")

    default_meta_rows = parse_meta_rows(args.meta_rows)
    schema = load_schema(args.schema) if args.schema else {"field_rows": {}, "meta_rows": {}, "refs": []}

    load_workbook = load_openpyxl()
    wb = load_workbook(args.workbook, data_only=False, keep_vba=args.workbook.suffix.lower() == ".xlsm")

    sheets: dict[str, dict[str, Any]] = {}
    for ws in wb.worksheets:
        if ws.title in schema["field_rows"]:
            field_row = schema["field_rows"][ws.title]
            meta_rows = schema["meta_rows"].get(ws.title, [])
        elif args.field_row is not None:
            field_row = args.field_row
            meta_rows = schema["meta_rows"].get(ws.title, default_meta_rows)
        else:
            # No explicit config — auto-detect per sheet.
            field_row, meta_rows = auto_detect_layout(ws)
        sheets[ws.title] = read_sheet(ws, field_row, meta_rows)

    refs: list[dict[str, str]] = []
    if not args.no_auto:
        refs.extend(auto_detect_refs(sheets))
    for ref in schema["refs"]:
        ref_copy = {**ref, "source": "schema"}
        # de-dup against auto-detected
        existing = next(
            (
                r for r in refs
                if r["from_sheet"] == ref_copy["from_sheet"]
                and r["from_field"] == ref_copy["from_field"]
                and r["to_sheet"] == ref_copy["to_sheet"]
                and r["to_field"] == ref_copy["to_field"]
            ),
            None,
        )
        if existing:
            existing["source"] = "schema+auto"
        else:
            refs.append(ref_copy)

    checked_refs = [check_ref(ref, sheets) for ref in refs]

    report = {
        "workbook": str(args.workbook),
        "sheets": [
            {
                "name": info["name"],
                "primary_key": info["primary_key"],
                "field_row": info["field_row"],
                "meta_rows": info["meta_rows"],
                "data_start_row": info["data_start_row"],
                "columns": {k: [(r, v) for r, v in cells] for k, cells in info["columns"].items()},
            }
            for info in sheets.values()
        ],
        "refs": checked_refs,
    }

    if args.format == "json":
        # strip per-row column data from JSON output to keep it small
        slim = {
            "workbook": report["workbook"],
            "sheets": [
                {k: v for k, v in s.items() if k != "columns"}
                for s in report["sheets"]
            ],
            "refs": checked_refs,
        }
        text = json.dumps(slim, ensure_ascii=False, indent=2)
    else:
        text = render_md(report)

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text, encoding="utf-8")
    else:
        sys.stdout.write(text)

    # Non-zero exit when there are orphans, so CI / scripts can detect failures.
    total_orphans = sum(r.get("orphan_count", 0) for r in checked_refs)
    if total_orphans > 0:
        sys.stderr.write(f"\n[validate_refs] {total_orphans} orphan(s) found across {len(checked_refs)} ref(s).\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
