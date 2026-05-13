#!/usr/bin/env python3
"""Scan a directory of config tables and emit an inventory.

Supports .xlsx, .xlsm, .csv, .tsv, .json.

Multi-row headers: pass --field-row N (1-based) to mark which row holds the
machine-readable field names, and --meta-rows R1,R2,... for extra header rows
(Chinese display name, type annotation, comment, etc.). If both are omitted the
script tries to detect a "row 1 Chinese / row 2 ASCII" pattern (common in game
configs) and recommends field-row=2 + meta-rows=1. Otherwise it falls back to
guessing a single-row header from the first 12 rows.

Encoding fallback: CSV/TSV/JSON files are first read as UTF-8 (with BOM tolerance);
on UnicodeDecodeError the script retries with GBK and records the encoding used
in the per-file output.

Files that look like patch JSON (`{"sheets": [{"sheet": ..., "updates": ...}]}`)
are skipped so a folder containing examples is not polluted.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import platform
import re
import sys
from pathlib import Path
from typing import Any, Iterable

# Sibling module — relies on the script's directory being on sys.path[0],
# which Python guarantees when invoked as `python3 scripts/foo.py`.
from _config_loader import check_paths, load_config_file, merge_into_args


SUPPORTED = {".xlsx", ".xlsm", ".csv", ".tsv", ".json"}

KEY_EXACT = {"id", "key", "编号", "主键", "uid"}
KEY_SUFFIX_RE = re.compile(r"_(id|key)$", re.IGNORECASE)
KEY_CAMEL_RE = re.compile(r"^[A-Za-z][A-Za-z0-9]*(ID|Id|Key)$")
CHINESE_RE = re.compile(r"[一-鿿]")
ASCII_FIELD_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]*$")
TYPE_ANNOTATION_VALUES = {
    "int", "integer", "str", "string", "float", "double", "bool", "boolean",
    "list", "array", "dict", "object", "json", "date", "datetime", "long",
    "short", "char", "varchar", "text", "number",
}
# Common Chinese keywords that show up in annotation/comment rows, not data rows.
CN_ANNOTATION_KEYWORDS = (
    "主键", "外键", "唯一", "必填", "可选", "注释", "说明", "类型", "枚举",
    "引用", "默认", "废弃", "弃用", "运行时", "客户端", "服务端",
)

ENCODING_FALLBACKS = ("utf-8-sig", "utf-8", "gbk", "gb18030")


def load_openpyxl():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required for Excel files.\n"
            "Install with:  pip install openpyxl  (or pip install -r requirements.txt)\n"
            "Or restrict --root to CSV/TSV/JSON files only."
        ) from exc
    return load_workbook


def tool_versions(include_openpyxl: bool = False) -> dict[str, str]:
    versions: dict[str, str] = {"python": platform.python_version()}
    if include_openpyxl:
        try:
            import openpyxl  # type: ignore
            versions["openpyxl"] = openpyxl.__version__
        except ModuleNotFoundError:
            pass
    return versions


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def trim(row: list[str]) -> list[str]:
    while row and row[-1] == "":
        row.pop()
    return row


def nonempty(row: list[str]) -> int:
    return sum(1 for value in row if value)


def is_key_candidate(name: str) -> bool:
    if not name:
        return False
    low = name.strip().lower()
    if low in KEY_EXACT:
        return True
    if KEY_SUFFIX_RE.search(name):
        return True
    if KEY_CAMEL_RE.match(name):
        return True
    return False


def looks_chinese_row(row: list[str]) -> bool:
    values = [v for v in row if v]
    if not values:
        return False
    chinese_count = sum(1 for v in values if CHINESE_RE.search(v))
    return chinese_count >= max(1, len(values) // 2)


def looks_ascii_field_row(row: list[str]) -> bool:
    values = [v for v in row if v]
    if not values:
        return False
    ascii_count = sum(1 for v in values if ASCII_FIELD_RE.match(v))
    return ascii_count >= max(1, len(values) * 2 // 3)


def looks_type_annotation_row(row: list[str]) -> bool:
    values = [v for v in row if v]
    if not values:
        return False
    type_hits = sum(1 for v in values if v.lower() in TYPE_ANNOTATION_VALUES)
    cn_hits = sum(
        1 for v in values
        if any(kw in v for kw in CN_ANNOTATION_KEYWORDS)
    )
    short_text_hits = sum(
        1 for v in values
        if len(v) <= 12 and not any(c.isdigit() for c in v) and CHINESE_RE.search(v)
    )
    return (type_hits + cn_hits) >= max(1, len(values) // 2) or short_text_hits == len(values)


def looks_data_row(row: list[str]) -> bool:
    """A row mostly composed of values that look like data (digits / paths / long text)."""
    values = [v for v in row if v]
    if not values:
        return False
    digit_heavy = sum(1 for v in values if any(c.isdigit() for c in v))
    return digit_heavy >= max(1, len(values) * 2 // 3)


def guess_header(rows: list[list[str]], scan_rows: int) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:scan_rows]):
        values = [value for value in row if value]
        if not values:
            continue
        unique = len(set(values))
        score = len(values) * 3 + unique - idx
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def field_summary(headers: list[str]) -> list[dict[str, Any]]:
    result = []
    for idx, name in enumerate(headers, start=1):
        if not name:
            continue
        result.append({"index": idx, "name": name, "key_candidate": is_key_candidate(name)})
    return result


def sample_rows(rows: list[list[str]], data_start_idx: int, headers: list[str], sample_count: int) -> list[dict[str, str]]:
    samples = []
    for row in rows[data_start_idx:]:
        if nonempty(row) == 0:
            continue
        item = {}
        for idx, name in enumerate(headers):
            if not name:
                continue
            value = row[idx] if idx < len(row) else ""
            if value:
                item[name] = value
        if item:
            samples.append(item)
        if len(samples) >= sample_count:
            break
    return samples


def is_patch_shaped_json(payload: Any) -> bool:
    """Detect JSON that looks like a patch_xlsx.py patch file, so we can skip it."""
    if not isinstance(payload, dict):
        return False
    sheets = payload.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        return False
    first = sheets[0]
    if not isinstance(first, dict):
        return False
    return "updates" in first or "appends" in first or "key_field" in first


def iter_config_files(root: Path, max_files: int, ignore: list[str]) -> Iterable[Path]:
    if root.is_file():
        if root.suffix.lower() in SUPPORTED:
            yield root
        return
    count = 0
    for path in sorted(root.rglob("*")):
        if not (path.is_file() and path.suffix.lower() in SUPPORTED):
            continue
        if any(path.match(pattern) or pattern in path.name for pattern in ignore):
            continue
        yield path
        count += 1
        if count >= max_files:
            return


def decode_with_fallback(data: bytes, prefer: str) -> tuple[str, str]:
    encodings = (prefer,) + tuple(e for e in ENCODING_FALLBACKS if e != prefer)
    last_err: UnicodeDecodeError | None = None
    for enc in encodings:
        try:
            return data.decode(enc), enc
        except UnicodeDecodeError as exc:
            last_err = exc
            continue
    raise last_err if last_err else UnicodeDecodeError(prefer, b"", 0, 1, "all encodings failed")


def detect_layout_hints(rows: list[list[str]], field_row_1based: int | None, meta_rows_1based: list[int]) -> list[str]:
    """Return human-readable warnings/hints about likely header misconfiguration."""
    hints: list[str] = []
    if not rows:
        return hints
    first = rows[0] if len(rows) > 0 else []
    second = rows[1] if len(rows) > 1 else []
    if field_row_1based is None and looks_chinese_row(first) and looks_ascii_field_row(second):
        hints.append(
            "row 1 looks like Chinese display names and row 2 looks like ASCII field names; "
            "consider --field-row 2 --meta-rows 1"
        )
    if field_row_1based is not None:
        field_idx = field_row_1based - 1
        if 0 <= field_idx < len(rows):
            field_row_content = rows[field_idx]
            if looks_data_row(field_row_content) and not looks_ascii_field_row(field_row_content):
                hints.append(
                    f"row {field_idx + 1} (your --field-row) looks like data, not headers — "
                    f"the field row may be different in this sheet"
                )
        data_start_idx = max([field_idx] + [r - 1 for r in meta_rows_1based]) + 1
        if 0 <= data_start_idx < len(rows):
            first_data = rows[data_start_idx]
            if looks_type_annotation_row(first_data):
                hints.append(
                    f"row {data_start_idx + 1} (the first row of data) looks like annotations / type "
                    f"comments ('int', '主键', etc.); did you mean to include it in --meta-rows?"
                )
    return hints


def auto_detect_layout(rows: list[list[str]]) -> tuple[int, list[int], int]:
    """Per-sheet auto-detection of (field_row, meta_rows, data_start_row), 1-based.

    Strategy:
    - Scan the first 8 rows.
    - Pick the row that "looks most like ASCII field names" (CamelCase / snake_case).
    - Above that row -> meta_rows. Annotation-looking rows immediately below -> meta_rows too.
    - data_start_row = max(field_row, *meta_rows) + 1.
    """
    scan = rows[:8]
    if not scan:
        return 1, [], 2
    best_field_row = 1
    best_score = -1
    for idx, row in enumerate(scan, start=1):
        if not looks_ascii_field_row(row):
            continue
        score = sum(1 for v in row if v)
        if score > best_score:
            best_score = score
            best_field_row = idx
    if best_score < 0:
        # No row looked like field names; fall back to guess.
        best_field_row = guess_header(rows, 8) + 1
    meta_rows = list(range(1, best_field_row))
    next_row = best_field_row + 1
    while next_row <= len(scan) and looks_type_annotation_row(scan[next_row - 1]):
        meta_rows.append(next_row)
        next_row += 1
    data_start = max([best_field_row] + meta_rows) + 1
    return best_field_row, sorted(meta_rows), data_start


def resolve_header_layout(
    rows: list[list[str]],
    field_row_1based: int | None,
    meta_rows_1based: list[int],
    header_scan_rows: int,
) -> tuple[int, int, list[str]]:
    if field_row_1based is not None:
        field_idx = field_row_1based - 1
        if field_idx < 0:
            field_idx = 0
        if field_idx >= len(rows):
            field_idx = max(len(rows) - 1, 0)
        headers = rows[field_idx] if field_idx < len(rows) else []
        last_header_row = max([field_idx] + [r - 1 for r in meta_rows_1based])
        data_start = last_header_row + 1
        return field_idx, data_start, headers
    field_idx = guess_header(rows, header_scan_rows)
    headers = rows[field_idx] if rows else []
    return field_idx, field_idx + 1, headers


def inspect_delimited(path: Path, delimiter: str, args: argparse.Namespace) -> dict[str, Any]:
    data = path.read_bytes()
    text, encoding = decode_with_fallback(data, args.encoding)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[list[str]] = []
    for idx, row in enumerate(reader):
        rows.append(trim([norm(cell) for cell in row[: args.max_cols]]))
        if idx + 1 >= args.max_scan_rows:
            break
    field_idx, data_start, headers = resolve_header_layout(
        rows, args.field_row, args.meta_rows, args.header_scan_rows
    )
    hints = detect_layout_hints(rows, args.field_row, args.meta_rows)
    return {
        "type": path.suffix.lower().lstrip("."),
        "path": str(path),
        "encoding": encoding,
        "hints": hints,
        "sheets": [
            {
                "name": path.stem,
                "observed_rows": len(rows),
                "observed_cols": max((len(row) for row in rows), default=0),
                "field_row": field_idx + 1,
                "meta_rows": list(args.meta_rows) if args.meta_rows else [],
                "data_start_row": data_start + 1,
                "fields": field_summary(headers),
                "samples": sample_rows(rows, data_start, headers, args.sample_rows),
            }
        ],
    }


def inspect_json(path: Path, args: argparse.Namespace) -> dict[str, Any] | None:
    data = path.read_bytes()
    text, encoding = decode_with_fallback(data, args.encoding)
    payload = json.loads(text)
    if is_patch_shaped_json(payload):
        return None
    rows = payload if isinstance(payload, list) else payload.get("rows", []) if isinstance(payload, dict) else []
    fields = []
    samples = []
    if rows and isinstance(rows[0], dict):
        names = sorted({key for row in rows[: args.sample_rows] for key in row.keys()})
        fields = field_summary(names)
        for row in rows[: args.sample_rows]:
            samples.append({str(k): norm(v) for k, v in row.items() if norm(v)})
    return {
        "type": "json",
        "path": str(path),
        "encoding": encoding,
        "hints": [],
        "sheets": [
            {
                "name": path.stem,
                "observed_rows": len(rows) if isinstance(rows, list) else 0,
                "observed_cols": len(fields),
                "field_row": None,
                "meta_rows": [],
                "data_start_row": None,
                "fields": fields,
                "samples": samples,
            }
        ],
    }


def inspect_excel(path: Path, args: argparse.Namespace) -> dict[str, Any]:
    load_workbook = load_openpyxl()
    wb = load_workbook(path, read_only=True, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    sheets = []
    file_hints: list[str] = []
    for ws in wb.worksheets[: args.max_sheets]:
        rows: list[list[str]] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(trim([norm(cell) for cell in row[: args.max_cols]]))
            if idx + 1 >= args.max_scan_rows:
                break
        field_idx, data_start, headers = resolve_header_layout(
            rows, args.field_row, args.meta_rows, args.header_scan_rows
        )
        sheet_hints = detect_layout_hints(rows, args.field_row, args.meta_rows)
        for hint in sheet_hints:
            file_hints.append(f"{ws.title}: {hint}")
        # Per-sheet auto-detect — used by --patch-template regardless of user CLI args,
        # since global CLI args don't fit heterogeneous-header workbooks.
        auto_field_row, auto_meta_rows, auto_data_start = auto_detect_layout(rows)
        auto_headers = rows[auto_field_row - 1] if rows and 0 <= auto_field_row - 1 < len(rows) else []
        sheets.append(
            {
                "name": ws.title,
                "declared_rows": ws.max_row,
                "declared_cols": ws.max_column,
                "observed_rows": len(rows),
                "observed_cols": max((len(row) for row in rows), default=0),
                "field_row": field_idx + 1,
                "meta_rows": list(args.meta_rows) if args.meta_rows else [],
                "data_start_row": data_start + 1,
                "fields": field_summary(headers),
                "samples": sample_rows(rows, data_start, headers, args.sample_rows),
                "_auto": {
                    "field_row": auto_field_row,
                    "meta_rows": auto_meta_rows,
                    "data_start_row": auto_data_start,
                    "fields": [f["name"] for f in field_summary(auto_headers)],
                    "key_field": next(
                        (f["name"] for f in field_summary(auto_headers) if f.get("key_candidate")),
                        None,
                    ),
                },
            }
        )
    return {
        "type": path.suffix.lower().lstrip("."),
        "path": str(path),
        "encoding": "binary",
        "hints": file_hints,
        "sheets": sheets,
    }


def inspect_file(path: Path, args: argparse.Namespace) -> dict[str, Any] | None:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return inspect_excel(path, args)
    if suffix == ".csv":
        return inspect_delimited(path, ",", args)
    if suffix == ".tsv":
        return inspect_delimited(path, "\t", args)
    if suffix == ".json":
        return inspect_json(path, args)
    raise SystemExit(f"Unsupported file: {path}")


def render_md(inventory: dict[str, Any]) -> str:
    versions = inventory.get("tool_versions") or {}
    versions_label = ", ".join(f"{k} {v}" for k, v in versions.items())
    lines = [
        "# Config Inventory",
        "",
        f"- Root: `{inventory['root']}`",
        f"- Files: {len(inventory['files'])}",
    ]
    if versions_label:
        lines.append(f"- Tool versions: {versions_label}")
    lines.append("")
    for file_info in inventory["files"]:
        lines.append(f"## {file_info['path']}")
        lines.append("")
        if file_info.get("encoding") and file_info["encoding"] != "binary":
            lines.append(f"- Encoding: `{file_info['encoding']}`")
        if file_info.get("hints"):
            for hint in file_info["hints"]:
                lines.append(f"- HINT: {hint}")
        if "error" in file_info:
            lines.append(f"- ERROR: {file_info['error']}")
            lines.append("")
            continue
        for sheet in file_info["sheets"]:
            size_rows = sheet.get("declared_rows", sheet.get("observed_rows"))
            size_cols = sheet.get("declared_cols", sheet.get("observed_cols"))
            field_row = sheet.get("field_row")
            meta_rows = sheet.get("meta_rows") or []
            data_start = sheet.get("data_start_row")
            header_info = f"field_row={field_row}"
            if meta_rows:
                header_info += f" meta_rows={','.join(str(r) for r in meta_rows)}"
            if data_start:
                header_info += f" data_start={data_start}"
            lines.append(f"- Sheet: `{sheet['name']}` rows={size_rows} cols={size_cols} {header_info}")
            keys = [f["name"] for f in sheet["fields"] if f["key_candidate"]]
            if keys:
                lines.append(f"  - Key candidates: {', '.join(keys)}")
            field_names = [f["name"] for f in sheet["fields"]]
            lines.append(f"  - Fields: {', '.join(field_names[:40])}")
            if sheet["samples"]:
                first = sheet["samples"][0]
                preview = " | ".join(f"{k}={v}" for k, v in list(first.items())[:8])
                lines.append(f"  - Sample: {preview}")
        lines.append("")
    return "\n".join(lines)


def parse_meta_rows(value: str) -> list[int]:
    if not value:
        return []
    parts = [p.strip() for p in value.split(",") if p.strip()]
    try:
        return [int(p) for p in parts]
    except ValueError as exc:
        raise SystemExit(f"--meta-rows expects comma-separated integers, got: {value}") from exc


def parse_ignore(value: str) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in value.split(",") if part.strip()]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect local configuration tables.")
    parser.add_argument(
        "--root", type=Path, default=None,
        help="Config root or single file. (Required, but can also be provided via --config.)",
    )
    parser.add_argument("--output", type=Path, help="Output path.")
    parser.add_argument("--format", choices=("json", "md"), default="json")
    parser.add_argument(
        "--field-row",
        type=int,
        default=None,
        help="1-based row index holding machine-readable field names. Omit to auto-guess.",
    )
    parser.add_argument(
        "--meta-rows",
        type=str,
        default="",
        help="Comma-separated 1-based row indices for supplementary header rows, e.g. 1,3,4.",
    )
    parser.add_argument(
        "--ignore",
        type=str,
        default="",
        help="Comma-separated filename patterns or substrings to skip, e.g. '*-patch.json,backup*'.",
    )
    parser.add_argument(
        "--patch-template",
        type=Path,
        default=None,
        help="Also write a pre-filled patch JSON skeleton to this path (one entry per "
             "Excel sheet, with detected field_row / meta_rows / data_start_row / key_field). "
             "Designed for the agent to fill in `updates` / `appends` without inventing the schema.",
    )
    parser.add_argument("--max-files", type=int, default=200)
    parser.add_argument("--max-sheets", type=int, default=80)
    parser.add_argument("--max-scan-rows", type=int, default=200)
    parser.add_argument("--header-scan-rows", type=int, default=12)
    parser.add_argument("--max-cols", type=int, default=80)
    parser.add_argument("--sample-rows", type=int, default=3)
    parser.add_argument(
        "--encoding",
        default="utf-8-sig",
        help="Preferred text encoding for CSV/TSV/JSON. Falls back to utf-8 / gbk / gb18030 on failure.",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help="Read all params from a UTF-8 JSON config file (bypasses argv encoding; "
             "use this on Windows when paths contain non-ASCII characters). "
             "Keys match the argparse dest names (e.g. 'field_row', 'meta_rows').",
    )
    args = parser.parse_args()
    # Merge --config FILE values onto args before any path is touched.
    cfg = load_config_file(args.config)
    merge_into_args(
        args, cfg,
        path_fields=("root", "output", "patch_template"),
        list_fields=("meta_rows", "ignore"),
    )
    # Type-coerce string forms passed via CLI (the JSON path is already typed).
    if isinstance(args.meta_rows, str):
        args.meta_rows = parse_meta_rows(args.meta_rows)
    if isinstance(args.ignore, str):
        args.ignore = parse_ignore(args.ignore)
    # Fail early with a friendly message if any path was mangled by argv encoding.
    check_paths(args, ("root", "output", "patch_template", "config"))
    return args


def build_patch_template(inventory: dict[str, Any]) -> dict[str, Any]:
    """Build a skeleton patch JSON the agent can fill in.

    One entry per Excel sheet (CSV/TSV/JSON files are skipped — patch_xlsx.py is
    Excel-only). Each entry uses **per-sheet auto-detection** (`_auto` block on
    each sheet dict) so heterogeneous-header workbooks come out correct, even if
    the inspect CLI was passed a single global --field-row / --meta-rows.
    """
    sheets: list[dict[str, Any]] = []
    for file_info in inventory["files"]:
        if file_info.get("type") not in {"xlsx", "xlsm"}:
            continue
        for sheet in file_info["sheets"]:
            auto = sheet.get("_auto") or {}
            entry: dict[str, Any] = {
                "sheet": sheet["name"],
                "field_row": auto.get("field_row") or sheet.get("field_row"),
                "data_start_row": auto.get("data_start_row") or sheet.get("data_start_row"),
                "updates": [],
                "appends": [],
                "_fields_available": auto.get("fields")
                or [f["name"] for f in sheet["fields"]],
            }
            meta_rows = auto.get("meta_rows") or []
            if meta_rows:
                entry["meta_rows"] = meta_rows
            key_field = auto.get("key_field") or next(
                (f["name"] for f in sheet["fields"] if f.get("key_candidate")),
                None,
            )
            if key_field:
                entry["key_field"] = key_field
            sheets.append(entry)
    return {
        "_generated_by": "inspect_config_tables.py --patch-template",
        "_help": (
            "Fill `updates` (cell tweaks by key/field or row/col) and/or `appends` "
            "(new rows by field name). See references/patch-format.md for the full spec. "
            "Keys starting with '_' are documentation only and ignored by patch_xlsx.py."
        ),
        "_root": inventory["root"],
        "sheets": sheets,
    }


def main() -> None:
    args = parse_args()
    if args.root is None:
        raise SystemExit("--root is required (pass it on the CLI or include it in --config).")
    if not args.root.exists():
        raise SystemExit(f"Root not found: {args.root}")
    files = []
    for path in iter_config_files(args.root, args.max_files, args.ignore):
        try:
            result = inspect_file(path, args)
            if result is None:
                continue  # skipped (e.g., patch-shaped JSON)
            files.append(result)
        except Exception as exc:
            files.append({"type": path.suffix.lower().lstrip("."), "path": str(path), "error": str(exc), "sheets": []})
    needs_openpyxl = any(f.get("type") in {"xlsx", "xlsm"} for f in files)
    inventory = {
        "root": str(args.root),
        "tool_versions": tool_versions(include_openpyxl=needs_openpyxl),
        "files": files,
    }
    text = json.dumps(inventory, ensure_ascii=False, indent=2) if args.format == "json" else render_md(inventory)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text, encoding="utf-8")
    else:
        sys.stdout.write(text)
    if args.patch_template:
        template = build_patch_template(inventory)
        args.patch_template.parent.mkdir(parents=True, exist_ok=True)
        args.patch_template.write_text(
            json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        sys.stderr.write(f"[inspect] patch template written to {args.patch_template}\n")


if __name__ == "__main__":
    main()
