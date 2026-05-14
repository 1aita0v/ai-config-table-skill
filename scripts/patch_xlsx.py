#!/usr/bin/env python3
"""Copy a workbook and apply structured updates / appends.

Supports multi-row headers via patch JSON fields `field_row` and `meta_rows`.
The legacy `header_row` field is still accepted (treated as `field_row`).

Always run with --dry-run first to preview cell-level changes before writing.

The source workbook is never modified. Output is written to --output. If patching
fails partway through, the partial output file is removed.

When --output already exists, by default a timestamped sibling is written
(e.g. `table_candidate_20260512_103045.xlsx`) so iterative debugging never
loses an existing candidate. Pass --force to overwrite, or --strict to error.
"""
from __future__ import annotations

import argparse
import datetime as dt
import difflib
import json
import platform
import re
import shutil
import sys
from pathlib import Path
from typing import Any

# Sibling module — relies on the script's directory being on sys.path[0].
from _config_loader import check_paths, load_config_file, merge_into_args


# Windows old-API path limit. Files written under longer paths can be silently
# created but unreadable by Excel / Explorer. Warn early.
LONG_PATH_LIMIT_HARD = 260
LONG_PATH_LIMIT_WARN = 240


def excel_lock_marker(path: Path) -> Path:
    """Path to the `~$X` lock marker Excel / WPS write next to an open workbook."""
    return path.with_name("~$" + path.name)


def assert_not_excel_locked(path: Path, label: str) -> None:
    """If a workbook is currently open in Excel / WPS, refuse and explain."""
    marker = excel_lock_marker(path)
    if marker.exists():
        raise SystemExit(
            f"{label} appears to be open in Excel / WPS: a lock marker exists at\n"
            f"  {marker}\n"
            f"Close the workbook in Excel / WPS first, then re-run."
        )


def warn_long_path(path: Path, label: str) -> None:
    text = str(path)
    if len(text) > LONG_PATH_LIMIT_WARN:
        msg = (
            f"[notice] --{label} is {len(text)} characters long. Older Windows "
            f"versions (or Excel) may refuse to open paths longer than "
            f"{LONG_PATH_LIMIT_HARD} chars. Consider a shorter path."
        )
        sys.stderr.write(msg + "\n")


def suggest(name: str, candidates: list[str], n: int = 3) -> str:
    """Format a 'did you mean ...?' suffix for an error message."""
    if not candidates:
        return ""
    matches = difflib.get_close_matches(name, candidates, n=n, cutoff=0.6)
    if not matches:
        return f"  Available: {', '.join(candidates[:10])}{' ...' if len(candidates) > 10 else ''}"
    return f"  Did you mean: {', '.join(matches)}?  (All: {', '.join(candidates[:10])}{' ...' if len(candidates) > 10 else ''})"


def load_openpyxl():
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import column_index_from_string, get_column_letter
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required for Excel patching.\n"
            "Install with:  pip install openpyxl"
        ) from exc
    return load_workbook, Font, PatternFill, column_index_from_string, get_column_letter


load_workbook, Font, PatternFill, column_index_from_string, get_column_letter = load_openpyxl()
MARK_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
COL_RE = re.compile(r"^[A-Z]{1,3}$")
FORMULA_ABORT_PREVIEW = 20

# Header names treated as the row-level note / comment column. First match wins.
NOTE_COLUMN_NAMES = ("备注", "注释", "Note", "Notes", "Comment", "Comments", "Remark", "Remarks")


def tool_versions() -> dict[str, str]:
    versions: dict[str, str] = {"python": platform.python_version()}
    try:
        import openpyxl  # type: ignore
        versions["openpyxl"] = openpyxl.__version__
    except ModuleNotFoundError:
        pass
    return versions


def detect_note_column(header: dict[str, int]) -> tuple[str | None, int | None]:
    """Return (header_name, col_index) of the first note-like column, or (None, None)."""
    for name in NOTE_COLUMN_NAMES:
        if name in header:
            return name, header[name]
    # Case-insensitive fallback for English names.
    lower_map = {k.lower(): k for k in header.keys()}
    for name in NOTE_COLUMN_NAMES:
        if name.lower() in lower_map:
            real = lower_map[name.lower()]
            return real, header[real]
    return None, None


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def one_line(value: Any, limit: int = 120) -> str:
    text = norm(value).replace("\n", "\\n")
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def cell_has_formula(cell) -> bool:
    value = cell.value
    return getattr(cell, "data_type", None) == "f" or (isinstance(value, str) and value.startswith("="))


def collect_formula_cells(wb, max_preview: int = FORMULA_ABORT_PREVIEW) -> dict[str, Any]:
    total = 0
    preview: list[dict[str, str]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not cell_has_formula(cell):
                    continue
                total += 1
                if len(preview) < max_preview:
                    preview.append(
                        {
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "formula": one_line(cell.value),
                        }
                    )
    return {"count": total, "preview": preview}


def format_formula_preview(formulas: dict[str, Any]) -> list[str]:
    lines = [
        "Formula preview:",
    ]
    for item in formulas["preview"]:
        lines.append(f"  {item['sheet']}!{item['cell']}: {item['formula']}")
    hidden = formulas["count"] - len(formulas["preview"])
    if hidden > 0:
        lines.append(f"  ... {hidden} more formula cell(s)")
    return lines


def assert_formula_free_workbook(wb, source: Path) -> None:
    formulas = collect_formula_cells(wb)
    if formulas["count"] == 0:
        return
    lines = [
        f"Source workbook contains {formulas['count']} formula cell(s): {source}",
        "Before patching config tables, ask the user how to handle formulas: clear/paste values, export a formula-free file, or explicitly preserve formulas with --allow-formulas.",
        "Why: formulas can calculate incorrectly after appending rows, and openpyxl cannot recalculate them.",
        *format_formula_preview(formulas),
    ]
    raise SystemExit("\n".join(lines))


def render_formula_override_warning(formulas: dict[str, Any], source: Path) -> list[str]:
    if formulas["count"] == 0:
        return []
    return [
        f"WARNING: --allow-formulas is set for {source}.",
        f"Source workbook still contains {formulas['count']} formula cell(s).",
        "Use this only when the user explicitly wants formulas preserved, accepts Excel/WPS recalculation risk, and has defined the expected formula results.",
        *format_formula_preview(formulas),
    ]


def resolve_col(col: str | int) -> int:
    if isinstance(col, int):
        return col
    if isinstance(col, str) and col.isdigit():
        return int(col)
    return column_index_from_string(str(col))


def mark(cell) -> None:
    cell.fill = MARK_FILL
    old = cell.font
    cell.font = Font(
        name=old.name,
        size=old.size,
        bold=old.bold,
        italic=old.italic,
        color="000000",
        underline="single",
    )


def row_values(ws, row_idx: int) -> list[str]:
    return [norm(ws.cell(row=row_idx, column=col).value) for col in range(1, ws.max_column + 1)]


def guess_field_row(ws, scan_rows: int = 12) -> int:
    best_row = 1
    best_score = -1
    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        row = row_values(ws, row_idx)
        values = [value for value in row if value]
        if not values:
            continue
        score = len(values) * 3 + len(set(values)) - row_idx
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def header_map(ws, field_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        name = norm(ws.cell(row=field_row, column=col).value)
        if name and name not in result:
            result[name] = col
    return result


def reverse_header_map(header: dict[str, int]) -> dict[int, str]:
    return {col: name for name, col in header.items()}


def build_key_index(ws, key_col: int, data_start: int) -> dict[str, int]:
    index: dict[str, int] = {}
    for row_idx in range(data_start, ws.max_row + 1):
        value = norm(ws.cell(row=row_idx, column=key_col).value)
        if value and value not in index:
            index[value] = row_idx
    return index


def last_meaningful_row(ws, min_row: int) -> int:
    last = min_row - 1
    for row_idx in range(min_row, ws.max_row + 1):
        if any(row_values(ws, row_idx)):
            last = row_idx
    return last


KNOWN_TOP_LEVEL_KEYS = {"sheets"}


def validate_patch_schema(patch: Any, patch_path: Path) -> None:
    """Reject obviously wrong patch shapes before they silently produce 0 updates.

    The only top-level key patch_xlsx.py reads is ``sheets``. A common AI
    mistake is to wrap everything in ``{"files": [{"path": ..., "sheets": [...]}]}``
    — that gets silently ignored (no ``sheets`` at top level → empty loop).
    Catch this loudly with a pointer to the canonical format.
    """
    if not isinstance(patch, dict):
        raise SystemExit(
            f"Patch file root must be a JSON object ({patch_path}); got {type(patch).__name__}. "
            f"Expected schema: {{'sheets': [...]}} — see references/patch-format.md."
        )
    if "sheets" not in patch:
        # Identify likely-wrong wrappers and call them out explicitly.
        wrong_wrappers = [k for k in ("files", "patches", "workbooks", "sheet") if k in patch]
        if wrong_wrappers:
            raise SystemExit(
                f"Patch file is missing top-level 'sheets' key ({patch_path}); "
                f"found {wrong_wrappers!r} instead. patch_xlsx.py expects a flat "
                f"{{'sheets': [...]}} — see references/patch-format.md. If you generated "
                f"this from inspect's --patch-template, the template uses 'sheets' at the "
                f"top level; don't wrap it in 'files'/'patches'/'workbooks'."
            )
        raise SystemExit(
            f"Patch file is missing top-level 'sheets' key ({patch_path}). "
            f"Expected schema: {{'sheets': [...]}} — see references/patch-format.md."
        )
    # Underscore-prefix keys are doc-only convention (inspect's --patch-template
    # writes _generated_by / _help / _root for human inspection); silently skip.
    unknown = {k for k in patch.keys() if k not in KNOWN_TOP_LEVEL_KEYS and not k.startswith("_")}
    if unknown:
        sys.stderr.write(
            f"[patch_xlsx] warning: unknown top-level keys in patch JSON: "
            f"{sorted(unknown)!r}. These will be ignored. Known keys: "
            f"{sorted(KNOWN_TOP_LEVEL_KEYS)!r}. See references/patch-format.md.\n"
        )


def resolve_sheet_layout(sheet_patch: dict[str, Any], ws) -> tuple[int, list[int], int]:
    field_row = sheet_patch.get("field_row") or sheet_patch.get("header_row")
    field_row = int(field_row) if field_row else guess_field_row(ws)
    meta_rows = sheet_patch.get("meta_rows") or []
    meta_rows = [int(r) for r in meta_rows]
    declared_data_start = sheet_patch.get("data_start_row")
    if declared_data_start:
        data_start = int(declared_data_start)
    else:
        last_header = max([field_row] + meta_rows)
        data_start = last_header + 1
    return field_row, meta_rows, data_start


def collect_planned_changes(ws, sheet_patch: dict[str, Any]) -> dict[str, Any]:
    """Compute (but do not apply) the changes a patch would make. Used by --dry-run."""
    field_row, meta_rows, data_start = resolve_sheet_layout(sheet_patch, ws)
    header = header_map(ws, field_row)
    col_to_name = reverse_header_map(header)
    key_field = sheet_patch.get("key_field")
    key_col = header.get(key_field) if key_field else None
    key_index = build_key_index(ws, key_col, data_start) if key_col else {}
    note_field, note_col = detect_note_column(header)

    planned_updates: list[dict[str, Any]] = []
    planned_note_writes: list[dict[str, Any]] = []
    for update in sheet_patch.get("updates", []):
        if "row" in update and "col" in update:
            row = int(update["row"])
            col = resolve_col(update["col"])
        else:
            if not key_field:
                raise SystemExit(f"Sheet {ws.title}: key-based update requires key_field")
            if "key" not in update or "field" not in update:
                raise SystemExit(f"Sheet {ws.title}: update requires row/col or key/field")
            key = norm(update["key"])
            if key not in key_index:
                raise SystemExit(
                    f"Key not found in {ws.title}: {key_field}={key}\n"
                    f"{suggest(key, sorted(key_index.keys()))}"
                )
            row = key_index[key]
            field = update["field"]
            if field not in header:
                raise SystemExit(
                    f"Field not found in {ws.title}: {field}\n"
                    f"{suggest(field, list(header.keys()))}"
                )
            col = header[field]
        before = norm(ws.cell(row=row, column=col).value)
        planned_updates.append(
            {
                "row": row,
                "col": col,
                "field": col_to_name.get(col, ""),
                "before": before,
                "after": norm(update.get("value")),
            }
        )
        note = update.get("note")
        if note and note_col is not None:
            before_note = norm(ws.cell(row=row, column=note_col).value)
            planned_note_writes.append(
                {
                    "row": row,
                    "col": note_col,
                    "field": note_field,
                    "before": before_note,
                    "after": norm(note),
                }
            )

    planned_appends: list[dict[str, Any]] = []
    append_start = last_meaningful_row(ws, data_start) + 1
    for row_offset, row_data in enumerate(sheet_patch.get("appends", [])):
        row_idx = append_start + row_offset
        cells: list[dict[str, Any]] = []
        for field, value in row_data.items():
            if field in header:
                col = header[field]
            elif str(field).isdigit() or COL_RE.match(str(field)):
                col = resolve_col(field)
            else:
                raise SystemExit(
                    f"Append field not found in {ws.title}: {field}\n"
                    f"{suggest(str(field), list(header.keys()))}"
                )
            cells.append({"col": col, "field": col_to_name.get(col, str(field)), "value": norm(value)})
        planned_appends.append({"row": row_idx, "cells": cells})

    return {
        "sheet": ws.title,
        "field_row": field_row,
        "meta_rows": meta_rows,
        "data_start_row": data_start,
        "note_column": note_field,
        "updates": planned_updates,
        "note_writes": planned_note_writes,
        "appends": planned_appends,
    }


def apply_sheet_patch(ws, sheet_patch: dict[str, Any], should_mark: bool) -> dict[str, int]:
    field_row, meta_rows, data_start = resolve_sheet_layout(sheet_patch, ws)
    header = header_map(ws, field_row)
    key_field = sheet_patch.get("key_field")
    key_col = header.get(key_field) if key_field else None
    key_index = build_key_index(ws, key_col, data_start) if key_col else {}
    _note_field, note_col = detect_note_column(header)

    changed = 0
    note_writes = 0
    appended = 0

    for update in sheet_patch.get("updates", []):
        if "row" in update and "col" in update:
            row = int(update["row"])
            col = resolve_col(update["col"])
        else:
            key = norm(update["key"])
            row = key_index[key]
            field = update["field"]
            col = header[field]
        cell = ws.cell(row=row, column=col)
        cell.value = update.get("value")
        if should_mark:
            mark(cell)
        changed += 1
        note = update.get("note")
        if note and note_col is not None:
            note_cell = ws.cell(row=row, column=note_col)
            note_cell.value = note
            if should_mark:
                mark(note_cell)
            note_writes += 1

    append_start = last_meaningful_row(ws, data_start) + 1
    for row_offset, row_data in enumerate(sheet_patch.get("appends", [])):
        row_idx = append_start + row_offset
        for field, value in row_data.items():
            if field in header:
                col = header[field]
            elif str(field).isdigit() or COL_RE.match(str(field)):
                col = resolve_col(field)
            else:
                raise SystemExit(
                    f"Append field not found in {ws.title}: {field}\n"
                    f"{suggest(str(field), list(header.keys()))}"
                )
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            if should_mark:
                mark(cell)
            changed += 1
        appended += 1

    return {"changed_cells": changed, "appended_rows": appended, "note_writes": note_writes}


def format_col(col: int, field: str | None) -> str:
    if field:
        return f"{col}({field})"
    return str(col)


def render_dry_run(
    plans: list[dict[str, Any]],
    formula_warning: dict[str, Any] | None = None,
    source: Path | None = None,
) -> str:
    lines = ["# Patch Dry Run", ""]
    if formula_warning and formula_warning["count"] and source:
        lines.extend(render_formula_override_warning(formula_warning, source))
        lines.append("")
    total_updates = total_appends = total_notes = 0
    for plan in plans:
        lines.append(f"## Sheet: {plan['sheet']}")
        layout = (
            f"  field_row={plan['field_row']} meta_rows={plan['meta_rows'] or '[]'} "
            f"data_start_row={plan['data_start_row']}"
        )
        if plan.get("note_column"):
            layout += f" note_column={plan['note_column']}"
        lines.append(layout)
        lines.append("")
        if plan["updates"]:
            lines.append("  Updates:")
            for u in plan["updates"]:
                before = u["before"] if u["before"] else "(empty)"
                col_label = format_col(u["col"], u.get("field"))
                lines.append(f"    row={u['row']} col={col_label}: {before}  =>  {u['after']}")
                total_updates += 1
            lines.append("")
        if plan.get("note_writes"):
            lines.append("  Notes:")
            for n in plan["note_writes"]:
                before = n["before"] if n["before"] else "(empty)"
                col_label = format_col(n["col"], n.get("field"))
                lines.append(f"    row={n['row']} col={col_label}: {before}  =>  {n['after']}")
                total_notes += 1
            lines.append("")
        if plan["appends"]:
            lines.append("  Appends:")
            for a in plan["appends"]:
                preview = ", ".join(
                    f"{c.get('field') or 'col' + str(c['col'])}={c['value']}" for c in a["cells"][:8]
                )
                lines.append(f"    row={a['row']}: {preview}")
                total_appends += 1
            lines.append("")
        if not plan["updates"] and not plan["appends"] and not plan.get("note_writes"):
            lines.append("  (no changes)")
            lines.append("")
    lines.append(
        f"Total: {total_updates} update(s), {total_notes} note write(s), "
        f"{total_appends} append row(s) across {len(plans)} sheet(s)."
    )
    lines.append("Dry run only — source workbook was not modified, no output file written.")
    return "\n".join(lines)


def write_with_fallback(source: Path, output: Path) -> tuple[Path, str | None]:
    """Copy source to output. If write fails, fall back to ~/Downloads.

    Returns (actual_path_written, optional_warning_message).
    """
    try:
        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, output)
        return output, None
    except PermissionError as exc:
        fallback_dir = Path.home() / "Downloads"
        try:
            fallback_dir.mkdir(parents=True, exist_ok=True)
            fallback = fallback_dir / output.name
            shutil.copy2(source, fallback)
            return (
                fallback,
                f"Could not write to {output} ({exc}). Wrote candidate to {fallback} instead.",
            )
        except Exception as fallback_exc:
            raise SystemExit(
                f"Could not write to {output} ({exc}) and fallback to {fallback_dir} also failed: {fallback_exc}"
            ) from fallback_exc


def auto_suffix_path(path: Path) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Copy and patch an Excel workbook.")
    parser.add_argument(
        "--source", type=Path, default=None,
        help="Source workbook. (Required, but can also be provided via --config.)",
    )
    parser.add_argument(
        "--output", type=Path, default=None,
        help="Candidate output path. (Required, but can also be provided via --config.)",
    )
    parser.add_argument(
        "--patch", type=Path, default=None,
        help="Patch JSON path. (Required, but can also be provided via --config.)",
    )
    parser.add_argument("--force", action="store_true", help="Overwrite existing output without renaming.")
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Error out if --output already exists (legacy behavior).",
    )
    parser.add_argument("--no-mark", action="store_true", help="Do not mark edited cells yellow.")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print planned changes without writing the output file.",
    )
    parser.add_argument(
        "--allow-formulas",
        action="store_true",
        help="Advanced override: patch even when the source workbook contains formulas. "
             "Use only after explicit user approval.",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help="Read all params from a UTF-8 JSON config file (use this on Windows when "
             "paths contain non-ASCII characters).",
    )
    args = parser.parse_args()
    cfg = load_config_file(args.config)
    merge_into_args(args, cfg, path_fields=("source", "output", "patch"))
    check_paths(args, ("source", "output", "patch", "config"))
    return args


def main() -> None:
    args = parse_args()
    for field in ("source", "output", "patch"):
        if getattr(args, field) is None:
            raise SystemExit(f"--{field} is required (pass it on the CLI or include it in --config).")
    if not args.source.exists():
        raise SystemExit(f"Source not found: {args.source}")
    if not args.patch.exists():
        raise SystemExit(f"Patch file not found: {args.patch}")
    # Excel / WPS leaves a `~$name.xlsx` marker when name.xlsx is open. Refuse
    # to overwrite a candidate that's currently open — saving would silently
    # fail or corrupt the file.
    if not args.dry_run:
        assert_not_excel_locked(args.output, "--output")
    warn_long_path(args.source, "source")
    warn_long_path(args.output, "output")
    if args.source.resolve() == args.output.resolve():
        raise SystemExit(
            f"--source and --output must be different files. "
            f"This script never edits the source in place; pick a candidate path like "
            f"{args.source.with_name(args.source.stem + '_candidate' + args.source.suffix).name}."
        )
    try:
        patch = json.loads(args.patch.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"Patch file is not valid JSON ({args.patch}): {exc}") from exc

    validate_patch_schema(patch, args.patch)

    if args.dry_run:
        wb = load_workbook(args.source, data_only=False, keep_vba=args.source.suffix.lower() == ".xlsm")
        try:
            formula_warning = None
            if not args.allow_formulas:
                assert_formula_free_workbook(wb, args.source)
            else:
                formula_warning = collect_formula_cells(wb)
            plans = []
            for sheet_patch in patch.get("sheets", []):
                sheet_name = sheet_patch["sheet"]
                if sheet_name not in wb.sheetnames:
                    raise SystemExit(
                        f"Sheet not found: {sheet_name}\n"
                        f"{suggest(sheet_name, list(wb.sheetnames))}"
                    )
                plans.append(collect_planned_changes(wb[sheet_name], sheet_patch))
            sys.stdout.write(render_dry_run(plans, formula_warning, args.source) + "\n")
            return
        finally:
            wb.close()

    # Validate against the source workbook BEFORE copying it to --output, so
    # bad sheet / key / field surfaces a friendly SystemExit (matching dry-run)
    # rather than a KeyError traceback after a half-written candidate exists.
    wb_validate = load_workbook(
        args.source, data_only=False, keep_vba=args.source.suffix.lower() == ".xlsm",
    )
    formula_warning = None
    try:
        if not args.allow_formulas:
            assert_formula_free_workbook(wb_validate, args.source)
        else:
            formula_warning = collect_formula_cells(wb_validate)
        for sheet_patch in patch.get("sheets", []):
            sheet_name = sheet_patch["sheet"]
            if sheet_name not in wb_validate.sheetnames:
                raise SystemExit(
                    f"Sheet not found: {sheet_name}\n"
                    f"{suggest(sheet_name, list(wb_validate.sheetnames))}"
                )
            collect_planned_changes(wb_validate[sheet_name], sheet_patch)
    finally:
        wb_validate.close()

    output_path = args.output
    warnings: list[str] = []
    if output_path.exists():
        if args.strict:
            raise SystemExit(f"Output already exists: {output_path}. (--strict was passed.)")
        if not args.force:
            new_path = auto_suffix_path(output_path)
            warnings.append(
                f"Output existed; writing to {new_path.name} instead. "
                f"Use --force to overwrite the existing file."
            )
            output_path = new_path

    output_path, fallback_warning = write_with_fallback(args.source, output_path)
    if fallback_warning:
        warnings.append(fallback_warning)

    wb = None
    try:
        wb = load_workbook(output_path, keep_vba=output_path.suffix.lower() == ".xlsm")
        summary = []
        for sheet_patch in patch.get("sheets", []):
            sheet_name = sheet_patch["sheet"]
            if sheet_name not in wb.sheetnames:
                raise SystemExit(
                    f"Sheet not found: {sheet_name}\n"
                    f"{suggest(sheet_name, list(wb.sheetnames))}"
                )
            result = apply_sheet_patch(wb[sheet_name], sheet_patch, not args.no_mark)
            result["sheet"] = sheet_name
            summary.append(result)
        wb.save(output_path)
    except BaseException:
        # Remove partially-written output so the user does not trust a broken file.
        try:
            output_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise
    finally:
        if wb is not None:
            wb.close()

    report: dict[str, Any] = {
        "output": str(output_path),
        "tool_versions": tool_versions(),
        "sheets": summary,
    }
    if warnings:
        report["warnings"] = warnings
        for w in warnings:
            sys.stderr.write(f"[notice] {w}\n")
    if formula_warning and formula_warning["count"]:
        report["formula_warning"] = formula_warning
        for line in render_formula_override_warning(formula_warning, args.source):
            sys.stderr.write(f"[notice] {line}\n")

    # 结束摘要 —— 告诉用户做了什么、副本在哪、下一步
    total_updates = sum(s.get("changed_cells", 0) for s in summary)
    total_appends = sum(s.get("appended_rows", 0) for s in summary)
    total_notes = sum(s.get("note_writes", 0) for s in summary)
    sys.stderr.write(
        "\n✅ 副本生成完成\n"
        f"  做了什么:{total_updates} 个单元格更新 / {total_appends} 个新行 / {total_notes} 个备注写入\n"
        f"  源表(没动):{args.source}\n"
        f"  副本(改了的格子标黄):{output_path}\n"
        "  下一步:① 跑 diff_config_tables.py 对比 / ② 跑 validate_refs.py 跨表对账 / ③ 都 OK 后让用户确认覆盖\n"
    )
    sys.stdout.write(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
