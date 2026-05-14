#!/usr/bin/env python3
"""Diff source and candidate Excel config files.

Compares cell values by default. Optionally compares merged-cell layouts with
--compare-merges. Other openpyxl-lossy attributes (data validation, conditional
formatting, embedded charts) are NOT compared — see validation-checklist.md for
manual checks.

When formula-bearing workbooks are intentionally allowed, use
--compare-formula-results after Excel / WPS / the project exporter recalculates
the candidate. This compares cached formula results (data_only=True); openpyxl
does not calculate formulas itself.

If the cell scan hits --max-cells, the report flags the diff as truncated
prominently at the top so callers don't trust a partial result.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

# Sibling module — relies on the script's directory being on sys.path[0].
from _config_loader import check_paths, load_config_file, merge_into_args


SUPPORTED = {".xlsx", ".xlsm"}


def load_openpyxl():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required for Excel diffing.\n"
            "Install with:  pip install openpyxl"
        ) from exc
    return load_workbook


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def compare_workbooks(
    source: Path,
    candidate: Path,
    max_cells: int,
    compare_merges: bool,
    compare_formula_results: bool,
) -> dict[str, Any]:
    load_workbook = load_openpyxl()
    src_wb = load_workbook(source, read_only=True, data_only=False, keep_vba=source.suffix.lower() == ".xlsm")
    cand_wb = load_workbook(candidate, read_only=True, data_only=False, keep_vba=candidate.suffix.lower() == ".xlsm")
    src_cached_wb = None
    cand_cached_wb = None
    try:
        if compare_formula_results:
            src_cached_wb = load_workbook(
                source, read_only=True, data_only=True, keep_vba=source.suffix.lower() == ".xlsm"
            )
            cand_cached_wb = load_workbook(
                candidate, read_only=True, data_only=True, keep_vba=candidate.suffix.lower() == ".xlsm"
            )
        src_sheets = set(src_wb.sheetnames)
        cand_sheets = set(cand_wb.sheetnames)
        result: dict[str, Any] = {
            "source": str(source),
            "candidate": str(candidate),
            "added_sheets": sorted(cand_sheets - src_sheets),
            "removed_sheets": sorted(src_sheets - cand_sheets),
            "sheets": [],
            "truncated": False,
            "compare_formula_results": compare_formula_results,
        }
        seen_cells = 0
        for sheet_name in src_wb.sheetnames:
            if sheet_name not in cand_sheets:
                continue
            src = src_wb[sheet_name]
            cand = cand_wb[sheet_name]
            src_cached = src_cached_wb[sheet_name] if src_cached_wb and sheet_name in src_cached_wb.sheetnames else None
            cand_cached = (
                cand_cached_wb[sheet_name] if cand_cached_wb and sheet_name in cand_cached_wb.sheetnames else None
            )
            max_row = max(src.max_row, cand.max_row)
            max_col = max(src.max_column, cand.max_column)
            changes = []
            formula_changes = []
            formula_result_changes = []
            missing_formula_results = []
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    seen_cells += 1
                    if seen_cells > max_cells:
                        result["truncated"] = True
                        break
                    left_raw = src.cell(row=row, column=col).value
                    right_raw = cand.cell(row=row, column=col).value
                    left = norm(left_raw)
                    right = norm(right_raw)
                    left_is_formula = is_formula(left_raw)
                    right_is_formula = is_formula(right_raw)
                    if left != right:
                        change = {"row": row, "col": col, "before": left, "after": right}
                        changes.append(change)
                        if left_is_formula or right_is_formula:
                            formula_changes.append(change)
                    if compare_formula_results and (left_is_formula or right_is_formula):
                        left_cached_raw = src_cached.cell(row=row, column=col).value if src_cached else None
                        right_cached_raw = cand_cached.cell(row=row, column=col).value if cand_cached else None
                        left_cached = norm(left_cached_raw)
                        right_cached = norm(right_cached_raw)
                        formula_result = {
                            "row": row,
                            "col": col,
                            "before": left_cached,
                            "after": right_cached,
                            "source_formula": left,
                            "candidate_formula": right,
                        }
                        if left_cached != right_cached:
                            formula_result_changes.append(formula_result)
                        if right_is_formula and right_cached_raw is None:
                            missing_formula_results.append(formula_result)
                if result["truncated"]:
                    break
            sheet_entry = {
                "name": sheet_name,
                "source_size": {"rows": src.max_row, "cols": src.max_column},
                "candidate_size": {"rows": cand.max_row, "cols": cand.max_column},
                "changed_cells": len(changes),
                "formula_changes": len(formula_changes),
                "changes_preview": changes[:50],
                "formula_changes_preview": formula_changes[:20],
            }
            if compare_formula_results:
                sheet_entry["formula_result_changes"] = len(formula_result_changes)
                sheet_entry["missing_formula_results"] = len(missing_formula_results)
                sheet_entry["formula_result_changes_preview"] = formula_result_changes[:20]
                sheet_entry["missing_formula_results_preview"] = missing_formula_results[:20]
            if compare_merges:
                # Reload non-read-only to access merged_cells.
                src_full = load_workbook(
                    source, read_only=False, data_only=False, keep_vba=source.suffix.lower() == ".xlsm"
                )
                cand_full = load_workbook(
                    candidate, read_only=False, data_only=False, keep_vba=candidate.suffix.lower() == ".xlsm"
                )
                try:
                    src_merges = (
                        {str(r) for r in src_full[sheet_name].merged_cells.ranges}
                        if sheet_name in src_full.sheetnames
                        else set()
                    )
                    cand_merges = (
                        {str(r) for r in cand_full[sheet_name].merged_cells.ranges}
                        if sheet_name in cand_full.sheetnames
                        else set()
                    )
                    sheet_entry["merge_added"] = sorted(cand_merges - src_merges)
                    sheet_entry["merge_removed"] = sorted(src_merges - cand_merges)
                finally:
                    src_full.close()
                    cand_full.close()
            result["sheets"].append(sheet_entry)
            if result["truncated"]:
                break
        return result
    finally:
        src_wb.close()
        cand_wb.close()
        if src_cached_wb:
            src_cached_wb.close()
        if cand_cached_wb:
            cand_cached_wb.close()


def workbook_pairs(source: Path, candidate: Path) -> list[tuple[Path, Path]]:
    if source.is_file() and candidate.is_file():
        return [(source, candidate)]
    pairs = []
    source_files = {p.relative_to(source): p for p in source.rglob("*") if p.suffix.lower() in SUPPORTED}
    candidate_files = {p.relative_to(candidate): p for p in candidate.rglob("*") if p.suffix.lower() in SUPPORTED}
    for rel in sorted(source_files.keys() & candidate_files.keys()):
        pairs.append((source_files[rel], candidate_files[rel]))
    return pairs


def render_md(diff: dict[str, Any]) -> str:
    lines = ["# Config Diff", ""]
    any_truncated = any(item.get("truncated") for item in diff["comparisons"])
    if any_truncated:
        lines.append("> **WARNING: diff was truncated by --max-cells. Results below are partial; raise --max-cells and re-run before trusting them.**")
        lines.append("")
    for item in diff["comparisons"]:
        lines.append(f"## {item['source']}")
        lines.append("")
        lines.append(f"- Candidate: `{item['candidate']}`")
        if item["added_sheets"]:
            lines.append(f"- Added sheets: {', '.join(item['added_sheets'])}")
        if item["removed_sheets"]:
            lines.append(f"- Removed sheets: {', '.join(item['removed_sheets'])}")
        if item["truncated"]:
            lines.append("- **Truncated**: cell scan stopped before completion.")
        lines.append("")
        if item.get("compare_formula_results"):
            lines.append(
                "| Sheet | Changed Cells | Formula Changes | Formula Result Changes | Missing Formula Results | Source Size | Candidate Size |"
            )
            lines.append("|---|---:|---:|---:|---:|---|---|")
        else:
            lines.append("| Sheet | Changed Cells | Formula Changes | Source Size | Candidate Size |")
            lines.append("|---|---:|---:|---|---|")
        for sheet in item["sheets"]:
            source_size = f"{sheet['source_size']['rows']}x{sheet['source_size']['cols']}"
            cand_size = f"{sheet['candidate_size']['rows']}x{sheet['candidate_size']['cols']}"
            if item.get("compare_formula_results"):
                lines.append(
                    f"| {sheet['name']} | {sheet['changed_cells']} | {sheet['formula_changes']} | "
                    f"{sheet.get('formula_result_changes', 0)} | {sheet.get('missing_formula_results', 0)} | "
                    f"{source_size} | {cand_size} |"
                )
            else:
                lines.append(
                    f"| {sheet['name']} | {sheet['changed_cells']} | {sheet['formula_changes']} | {source_size} | {cand_size} |"
                )
        lines.append("")
        for sheet in item["sheets"]:
            if sheet["changes_preview"]:
                lines.append(f"### {sheet['name']} Changes Preview")
                lines.append("")
                lines.append("| Row | Col | Before | After |")
                lines.append("|---:|---:|---|---|")
                for change in sheet["changes_preview"]:
                    before = change["before"].replace("|", "\\|").replace("\n", "\\n")
                    after = change["after"].replace("|", "\\|").replace("\n", "\\n")
                    lines.append(f"| {change['row']} | {change['col']} | {before} | {after} |")
                lines.append("")
            if sheet.get("formula_result_changes_preview"):
                lines.append(f"### {sheet['name']} Formula Result Changes Preview")
                lines.append("")
                lines.append("| Row | Col | Before Cached Result | After Cached Result | Source Formula | Candidate Formula |")
                lines.append("|---:|---:|---|---|---|---|")
                for change in sheet["formula_result_changes_preview"]:
                    before = change["before"].replace("|", "\\|").replace("\n", "\\n")
                    after = change["after"].replace("|", "\\|").replace("\n", "\\n")
                    source_formula = change["source_formula"].replace("|", "\\|").replace("\n", "\\n")
                    candidate_formula = change["candidate_formula"].replace("|", "\\|").replace("\n", "\\n")
                    lines.append(
                        f"| {change['row']} | {change['col']} | {before} | {after} | "
                        f"{source_formula} | {candidate_formula} |"
                    )
                lines.append("")
            if sheet.get("missing_formula_results_preview"):
                lines.append(f"### {sheet['name']} Missing Formula Cached Results")
                lines.append("")
                lines.append(
                    "- Candidate has formula cells without cached calculated results. "
                    "Open/recalculate the candidate in Excel/WPS or the project export tool, save it, then rerun diff."
                )
                lines.append("")
                lines.append("| Row | Col | Candidate Formula |")
                lines.append("|---:|---:|---|")
                for change in sheet["missing_formula_results_preview"]:
                    candidate_formula = change["candidate_formula"].replace("|", "\\|").replace("\n", "\\n")
                    lines.append(f"| {change['row']} | {change['col']} | {candidate_formula} |")
                lines.append("")
            if sheet.get("merge_added") or sheet.get("merge_removed"):
                lines.append(f"### {sheet['name']} Merge Changes")
                if sheet.get("merge_added"):
                    lines.append(f"- Added merges: {', '.join(sheet['merge_added'])}")
                if sheet.get("merge_removed"):
                    lines.append(f"- Removed merges: {', '.join(sheet['merge_removed'])}")
                lines.append("")
    if any_truncated:
        lines.append("> Reminder: this diff is truncated. Verify completeness by raising --max-cells.")
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Diff source and candidate Excel config files.")
    parser.add_argument(
        "--source", type=Path, default=None,
        help="Source workbook. (Required, but can also be provided via --config.)",
    )
    parser.add_argument(
        "--candidate", type=Path, default=None,
        help="Candidate workbook. (Required, but can also be provided via --config.)",
    )
    parser.add_argument("--output", type=Path)
    parser.add_argument("--format", choices=("md", "json"), default="md")
    parser.add_argument("--max-cells", type=int, default=200000)
    parser.add_argument(
        "--compare-merges",
        action="store_true",
        help="Also compare merged-cell ranges (slower; reloads workbooks without read_only).",
    )
    parser.add_argument(
        "--compare-formula-results",
        action="store_true",
        help="Also compare cached formula calculation results. Run after Excel/WPS/project tools recalculate the candidate.",
    )
    parser.add_argument(
        "--config", type=Path, default=None,
        help="Read all params from a UTF-8 JSON config file (use this on Windows when "
             "paths contain non-ASCII characters).",
    )
    args = parser.parse_args()
    cfg = load_config_file(args.config)
    merge_into_args(args, cfg, path_fields=("source", "candidate", "output"))
    check_paths(args, ("source", "candidate", "output", "config"))
    return args


def main() -> None:
    args = parse_args()
    for field in ("source", "candidate"):
        if getattr(args, field) is None:
            raise SystemExit(f"--{field} is required (pass it on the CLI or include it in --config).")
    if not args.source.exists():
        raise SystemExit(f"Source not found: {args.source}")
    if not args.candidate.exists():
        raise SystemExit(f"Candidate not found: {args.candidate}")
    comparisons = [
        compare_workbooks(src, cand, args.max_cells, args.compare_merges, args.compare_formula_results)
        for src, cand in workbook_pairs(args.source, args.candidate)
    ]
    diff = {"comparisons": comparisons}
    text = json.dumps(diff, ensure_ascii=False, indent=2) if args.format == "json" else render_md(diff)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text, encoding="utf-8")
    else:
        sys.stdout.write(text)


if __name__ == "__main__":
    main()
